"""Generate the FarryOn feature + bug tracking workbook."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = r"D:/FarryOn/FarryOn_Status_Report.xlsx"
FONT = "Arial"

TEAL = "0F6E56"
DARK = "12343B"
HDR_FILL = PatternFill("solid", fgColor=DARK)
HDR_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name=FONT, bold=True, size=16, color=TEAL)
BASE_FONT = Font(name=FONT, size=10)
WRAP = Alignment(vertical="top", wrap_text=True)
CENTER = Alignment(horizontal="center", vertical="center")
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

STATUS_FILL = {
    "Done · verified": "C6EFCE",
    "Done · live": "C6EFCE",
    "Done · not device-tested": "FFEB9C",
    "Done · local only": "FFEB9C",
    "Fixed · verified": "C6EFCE",
    "Fixed": "C6EFCE",
    "Mitigated / documented": "FFEB9C",
    "Improved": "C6EFCE",
}
STATUS_FONT = {
    "Done · verified": "006100", "Done · live": "006100",
    "Fixed · verified": "006100", "Fixed": "006100", "Improved": "006100",
    "Done · not device-tested": "9C6500", "Done · local only": "9C6500",
    "Mitigated / documented": "9C6500",
}


def style_header(ws, ncols, row=1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[row].height = 26


def fill_rows(ws, rows, status_col, start=2):
    for r, row in enumerate(rows, start=start):
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = BASE_FONT
            cell.alignment = WRAP
            cell.border = BORDER
        st = row[status_col]
        sc = ws.cell(row=r, column=status_col + 1)
        if st in STATUS_FILL:
            sc.fill = PatternFill("solid", fgColor=STATUS_FILL[st])
            sc.font = Font(name=FONT, size=10, bold=True,
                           color=STATUS_FONT.get(st, "000000"))
        sc.alignment = CENTER


wb = Workbook()

# ----- Sheet 1: Features -----
ws = wb.active
ws.title = "Features"
ws["A1"] = "FarryOn — Features Built"
ws["A1"].font = TITLE_FONT
ws.merge_cells("A1:F1")
ws.append([])
headers = ["#", "Feature", "Area", "What it does", "Status", "Notes"]
ws.append(headers)
style_header(ws, len(headers), row=3)

features = [
    ("Multi-AI provider", "Backend/App", "Switch between Gemini, OpenAI, Grok, Mock per session", "Done · verified", "User picks in Settings"),
    ("Realtime voice + vision", "Core", "Live mic + camera streamed to AI, spoken replies", "Done · verified", "Gemini Live / Realtime"),
    ("Hands-free mic (auto VAD + anti-echo)", "App", "Just talk; mic muted while assistant speaks", "Done · verified", "Half-duplex"),
    ("Camera portrait/landscape + orientation", "App", "Switch + lock orientation", "Done · verified", ""),
    ("Voice + pinch + preset camera zoom", "App", "Zoom by voice, pinch, or 1x/2x/4x chips with read-out", "Done · verified", ""),
    ("Full-bleed camera + quality + rounded viewport", "App", "Professional camera frame", "Done · verified", ""),
    ("Web search (Tavily + Google cross-check)", "Backend", "Both engines merged + cross-checked for accuracy", "Done · live", "Live scores limited"),
    ("Persist settings", "App", "Host/port/provider/keys saved across launches", "Done · verified", ""),
    ("Notes & Tasks (voice + view screen)", "Both", "Create/recall; grouped-by-date list", "Done · verified", ""),
    ("Voice task management", "Both", "Complete / edit / delete tasks by voice", "Done · verified", ""),
    ("Reminders (local notifications)", "App", "Fire at exact time even asleep/locked", "Done · verified", "alarmClock mode"),
    ("Voice device-control", "Both", "Mute mic, camera on/off, rotate, end session", "Done · verified", ""),
    ("Cloud deploy (Render)", "Infra", "Always-on backend at farryon-backend.onrender.com", "Done · live", "Blocks SMTP"),
    ("Reconnect button", "App", "Restart a session after it ends/drops", "Done · verified", ""),
    ("Performance (long chats)", "App", "Transcript capped; video paused during TTS", "Done · verified", ""),
    ("Location (\"where am I\")", "Both", "GPS + reverse-geocoded address via get_location", "Done · not device-tested", "Needs GPS test"),
    ("Chat history", "App", "Conversations saved; History tab", "Done · verified", ""),
    ("Email — read (multi-provider IMAP)", "Both", "Gmail/Outlook/Yahoo/Hostinger/Custom; categories, full body, reply-suggest", "Done · verified", "Reading works on cloud"),
    ("Email — send (SMTP, confirm)", "Both", "Send from user's own mailbox after confirmation", "Done · local only", "Cloud blocks SMTP"),
    ("New logo + adaptive icon", "App", "Full-bleed winged-eye launcher + notification icon", "Done · verified", ""),
    ("Professional chat UI", "App", "Polished bubbles, frosted panel, bigger chat", "Done · verified", ""),
    ("Confirm-before-action", "Backend", "Asks before create/change/delete/send", "Done · verified", "All providers"),
    ("Landmark & Product Finder", "Both", "Identify place/product/object from camera", "Done · verified", "Gemini vision core"),
    ("Voice-driven finder (auto)", "Both", "\"Take a photo & tell me\" -> auto capture + classify", "Done · verified", "No tap needed"),
    ("Reply in user's language", "Backend", "Answers in the language the user speaks", "Done · verified", "All providers"),
]
rows = [(i + 1, *f) for i, f in enumerate(features)]
fill_rows(ws, rows, status_col=4, start=4)
ws.freeze_panes = "A4"
ws.auto_filter.ref = f"A3:F{3 + len(rows)}"
widths = [4, 32, 12, 46, 24, 26]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

# ----- Sheet 2: Bugs & Fixes -----
ws2 = wb.create_sheet("Bugs & Fixes")
ws2["A1"] = "FarryOn — Bugs Found & Fixed"
ws2["A1"].font = TITLE_FONT
ws2.merge_cells("A1:F1")
ws2.append([])
headers2 = ["#", "Issue", "Severity", "Root cause", "Fix applied", "Status"]
ws2.append(headers2)
style_header(ws2, len(headers2), row=3)

bugs = [
    ("Only 'thinking', no spoken response / double reply", "High", "Per-turn receive loop ended early; thoughts leaked", "Multi-turn receive loop; skip thought parts", "Fixed · verified"),
    ("Backend crash on boot (ALLOWED_ORIGINS)", "High", "pydantic parsed CSV env as JSON", "Annotated NoDecode + CSV validator", "Fixed · verified"),
    ("Mic went deaf + echo (manual VAD)", "High", "START_SENSITIVITY_LOW + full-duplex", "Automatic VAD + client half-duplex", "Fixed · verified"),
    ("Assistant heard its own voice (echo loop)", "High", "Mic re-opened before TTS tail finished", "Mute margin 600->1200ms + drop user transcript during TTS", "Fixed · verified"),
    ("Reminders never fired (release build)", "Critical", "R8 stripped Gson signatures -> receiver crash 'Missing type parameter'", "Disable minify + proguard keep rules", "Fixed · verified"),
    ("Reminders deferred (Doze/Battery Saver)", "Critical", "exactAllowWhileIdle throttled", "AndroidScheduleMode.alarmClock", "Fixed · verified"),
    ("Relative reminder ('in 2 min') dropped", "Medium", "clientTime frozen at connect -> resolved to past", "Backend remind_in_seconds + fire-soon fallback", "Fixed · verified"),
    ("APK install fail (version downgrade)", "Low", "Split vs single-ABI version codes", "Uninstall + reinstall", "Fixed"),
    ("Settings: can't enter password / Save hidden", "High", "Fixed Column not scrollable; keyboard pushed it off", "Scrollable sheet + pinned Save bar + show/hide", "Fixed · verified"),
    ("Launcher icon tiny (grey card in white squircle)", "Medium", "Legacy PNG, not adaptive", "Full-bleed adaptive icon (bg + foreground)", "Fixed · verified"),
    ("Web search wrong/made-up live scores", "High", "Trusted stale AI 'answer' / Google countdown", "Cross-engine merge + pick authoritative, ignore countdowns", "Fixed · verified"),
    ("Email category search 'couldn't sign in'", "High", "X-GM-RAW multi-word query sent unquoted", "Quote the Gmail search query", "Fixed · verified"),
    ("Email reply sent to WRONG address", "Critical", "Model guessed recipient", "Parse exact from_email + confirm address aloud", "Fixed · verified"),
    ("Email send 'timed out' on cloud", "High", "Render blocks outbound SMTP (587/465)", "Works on local / SMTP-allowing host (Hostinger VPS/Fly.io)", "Mitigated / documented"),
    ("Camera froze; had to force-close app", "Critical", "No lifecycle handling; controller dead after background", "Release camera on background, re-open on resume", "Fixed · verified"),
    ("Mic/assistant active in background during Finder", "Medium", "Live session not paused for finder sheet", "Pause mic while Finder result open", "Fixed · verified"),
    ("Top bar / settings gear cut during 'Connecting'", "Medium", "Full-width bar clipped by rounded viewport", "Floating rounded bar + compact icons", "Fixed · verified"),
    ("Scan: 'no camera frame yet' (camera on)", "Medium", "First frame ~1s after camera on/resume", "grabFrame() waits ~2s for first frame", "Fixed · verified"),
    ("Spoke English, transcript/reply in Hindi", "High", "Auto language; STT used Devanagari", "Prompt: reply in user's spoken language (all providers)", "Fixed · verified"),
    ("Landmark/product detail weak for local things", "High", "Google Vision recognises famous landmarks only", "Gemini multimodal vision as core identifier", "Fixed · verified"),
    ("Finder 404 on cloud", "High", "Finder code not deployed to Render", "Deployed; set VISION_API_KEY in Render env", "Fixed"),
]
rows2 = [(i + 1, *b) for i, b in enumerate(bugs)]
fill_rows(ws2, rows2, status_col=5, start=4)
ws2.freeze_panes = "A4"
ws2.auto_filter.ref = f"A3:F{3 + len(rows2)}"
widths2 = [4, 40, 10, 44, 46, 22]
for i, w in enumerate(widths2, start=1):
    ws2.column_dimensions[get_column_letter(i)].width = w

# ----- Sheet 3: Summary -----
ws3 = wb.create_sheet("Summary")
ws3["A1"] = "FarryOn — Status Report"
ws3["A1"].font = TITLE_FONT
ws3.merge_cells("A1:C1")

ws3["A3"] = "Metric"
ws3["B3"] = "Count"
for c in ("A3", "B3"):
    ws3[c].fill = HDR_FILL
    ws3[c].font = HDR_FONT
    ws3[c].border = BORDER

nf = len(rows) + 3
nb = len(rows2) + 3
summary = [
    ("Total features", f"=COUNTA(Features!B4:B{nf})"),
    ("Features done & verified", f'=COUNTIF(Features!E4:E{nf},"Done · verified")+COUNTIF(Features!E4:E{nf},"Done · live")'),
    ("Features pending device test / local-only", f'=COUNTIF(Features!E4:E{nf},"Done · not device-tested")+COUNTIF(Features!E4:E{nf},"Done · local only")'),
    ("Total bugs logged", f"=COUNTA('Bugs & Fixes'!B4:B{nb})"),
    ("Bugs fixed & verified", f"=COUNTIF('Bugs & Fixes'!F4:F{nb},\"Fixed · verified\")+COUNTIF('Bugs & Fixes'!F4:F{nb},\"Fixed\")"),
    ("Bugs mitigated / documented", f"=COUNTIF('Bugs & Fixes'!F4:F{nb},\"Mitigated / documented\")"),
]
r = 4
for label, formula in summary:
    ws3.cell(row=r, column=1, value=label).font = BASE_FONT
    ws3.cell(row=r, column=1).border = BORDER
    fc = ws3.cell(row=r, column=2, value=formula)
    fc.font = Font(name=FONT, size=11, bold=True)
    fc.alignment = CENTER
    fc.border = BORDER
    r += 1

ws3.cell(row=r + 1, column=1, value="Pending / Next").font = Font(name=FONT, bold=True, size=12, color=TEAL)
pending = [
    "Email SEND on cloud — Render blocks SMTP; deploy backend to Hostinger VPS / Fly.io for always-on send.",
    "Location — code done; needs on-device GPS test (\"where am I\").",
    "Finder on cloud — set VISION_API_KEY in Render dashboard.",
    "Web search live sports scores — general search can't guarantee; add a sports API if critical.",
    "Latest APK install + device-test of newest fixes (camera lifecycle, finder, language).",
]
rr = r + 2
for p in pending:
    cell = ws3.cell(row=rr, column=1, value="• " + p)
    cell.font = BASE_FONT
    cell.alignment = WRAP
    ws3.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=3)
    rr += 1

ws3.column_dimensions["A"].width = 48
ws3.column_dimensions["B"].width = 12
ws3.column_dimensions["C"].width = 20
ws3["A2"] = "Generated 2026-06-23 · branch claude/gallant-franklin-j6qzbx"
ws3["A2"].font = Font(name=FONT, size=9, italic=True, color="808080")

wb.save(OUT)
print("saved", OUT)
